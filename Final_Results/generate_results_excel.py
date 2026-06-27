#!/usr/bin/env python3
"""Generate formatted Excel workbooks summarizing SAAF experiment results."""

from __future__ import annotations

import json
import statistics
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
PROJECT = ROOT.parent.parent


def load_json(rel_or_abs: str | Path) -> dict:
    path = Path(rel_or_abs)
    if not path.is_absolute():
        path = PROJECT / path
    with path.open() as f:
        return json.load(f)


def style_header(ws, row: int, ncols: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
    )
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def autosize_columns(ws, min_width: int = 10, max_width: int = 42) -> None:
    for col_cells in ws.columns:
        col_idx = col_cells[0].column
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(max_len + 2, min_width), max_width
        )


def write_table(ws, title: str, headers: list[str], rows: list[list], start_row: int = 1) -> int:
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
    header_row = start_row + 1
    for c, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=c, value=h)
    style_header(ws, header_row, len(headers))
    data_start = header_row + 1
    for r, row in enumerate(rows, data_start):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    autosize_columns(ws)
    ws.freeze_panes = ws.cell(row=data_start, column=1)
    return data_start + len(rows)


def pct(v: float, digits: int = 2) -> float:
    return round(v * 100, digits)


def build_final_aucroc() -> tuple[list[str], list[list]]:
    data = load_json(ROOT / "final_aucroc.json")
    headers = [
        "Dataset",
        "Videos",
        "Total Frames",
        "Positive Frames",
        "AUC-ROC (%)",
        "AUC-ROC",
    ]
    rows = []
    for r in data["results"]:
        rows.append(
            [
                r["dataset"],
                r["n_videos"],
                r["n_frames"],
                r["n_positives"],
                r["aucroc_pct"],
                round(r["aucroc"], 4),
            ]
        )
    return headers, rows


def build_stage_breakdown() -> tuple[list[str], list[list]]:
    data = load_json("ablation_saaf_breakdown/breakdown_results.json")
    headers = [
        "Dataset",
        "Stage 1 AUC-ROC (%)",
        "Stage 1 AP (%)",
        "Stage 2 AUC-ROC (%)",
        "Stage 2 AP (%)",
        "Stage 3 AUC-ROC (%)",
        "Stage 3 AP (%)",
        "Refine σ",
    ]
    rows = []
    for ds, stages in data["results"].items():
        rows.append(
            [
                ds,
                pct(stages["stage1_frame"]["auc_roc"]),
                pct(stages["stage1_frame"]["ap"]),
                pct(stages["stage2_gating"]["auc_roc"]),
                pct(stages["stage2_gating"]["ap"]),
                pct(stages["stage3_refined"]["auc_roc"]),
                pct(stages["stage3_refined"]["ap"]),
                stages["refine_sigma"],
            ]
        )
    return headers, rows


def build_fusion_ablation() -> tuple[list[str], list[list]]:
    datasets = [
        ("XD-Violence", "ablation_fusion_weight_XD_fill03_tf04/fusion_weight_results.json"),
        ("UCF-Crime", "ablation_fusion_weight_UCF_fill03_tf04/fusion_weight_results.json"),
        ("MSAD", "ablation_fusion_weight_MSAD_fill03_tf04/fusion_weight_results.json"),
        ("ShanghaiTech", "ablation_fusion_weight_ShanghaiTech_fill03_tf04/fusion_weight_results.json"),
    ]
    config_order = ["uniform", "mild", "moderate", "baseline", "strong", "reversed"]
    by_config: dict[str, dict] = {c: {} for c in config_order}
    lambdas: dict[str, tuple] = {}

    for ds_name, rel in datasets:
        data = load_json(rel)
        for item in data["results"]:
            cfg = item["config"]
            by_config.setdefault(cfg, {})[ds_name] = pct(item["auc_roc"])
            if cfg not in lambdas:
                lambdas[cfg] = (item["lambda_delta"], item["lambda_drift"], item["lambda_both"])

    headers = [
        "Config",
        "λ_δ",
        "λ_ρ",
        "λ_both",
        "XD AUC-ROC (%)",
        "UCF AUC-ROC (%)",
        "MSAD AUC-ROC (%)",
        "Shanghai AUC-ROC (%)",
    ]
    rows = []
    for cfg in config_order:
        ld, lr, lb = lambdas.get(cfg, ("", "", ""))
        rows.append(
            [
                cfg,
                ld,
                lr,
                lb,
                by_config[cfg].get("XD-Violence", ""),
                by_config[cfg].get("UCF-Crime", ""),
                by_config[cfg].get("MSAD", ""),
                by_config[cfg].get("ShanghaiTech", ""),
            ]
        )
    return headers, rows


def build_variance_seeds() -> tuple[list[str], list[list]]:
    data = load_json("0_Varienance_Mean/outpute_concat_scores/aucroc_sanity_check.json")
    by_ds: dict[str, list[float]] = {}
    detail_rows: list[list] = []
    for r in data["results"]:
        by_ds.setdefault(r["dataset"], []).append(r["aucroc"])
        detail_rows.append(
            [
                r["dataset"],
                r["seed"],
                pct(r["aucroc"]),
                round(r["aucroc"], 4),
                r["n_videos"],
                r["n_gt_frames"],
            ]
        )

    summary_headers = ["Dataset", "Mean AUC-ROC (%)", "Std (%)", "Seeds"]
    summary_rows = []
    for ds, scores in sorted(by_ds.items()):
        mean = statistics.mean(scores)
        std = statistics.stdev(scores) if len(scores) > 1 else 0.0
        summary_rows.append([ds, pct(mean), pct(std, 3), len(scores)])

    detail_headers = ["Dataset", "Seed", "AUC-ROC (%)", "AUC-ROC", "Videos", "GT Frames"]
    return summary_headers, summary_rows, detail_headers, detail_rows


def build_signal_attribution() -> tuple[list[str], list[list]]:
    data = load_json("ablation_saaf_breakdown/signal_attribution_distribution.json")
    headers = [
        "Dataset",
        "Videos",
        "Full-Res Frames",
        "Stride-Sampled Frames",
        "% Stride of Full",
        "Selected Frames",
        "% Selected of Stride",
        "Δ-only",
        "ρ-only",
        "Both",
        "% Δ-only",
        "% ρ-only",
        "% Both",
    ]
    rows = []
    for d in data["per_dataset"]:
        pct_sel = round(100 * d["selected_frames"] / d["stride_sampled_frames"], 4) if d["stride_sampled_frames"] else 0
        rows.append(
            [
                d["dataset"],
                d["json_files"],
                d["full_resolution_frames"],
                d["stride_sampled_frames"],
                d["pct_stride_of_full"],
                d["selected_frames"],
                pct_sel,
                d["delta_only"],
                d["drift_only"],
                d["both"],
                d["pct_delta_only_of_selected"],
                d["pct_drift_only_of_selected"],
                d["pct_both_of_selected"],
            ]
        )
    g = data["global_pooled"]
    rows.append(
        [
            "ALL (pooled)",
            "",
            g["full_resolution_frames"],
            g["stride_sampled_frames"],
            round(100 * g["stride_sampled_frames"] / g["full_resolution_frames"], 4),
            g["selected_frames"],
            round(100 * g["selected_frames"] / g["stride_sampled_frames"], 4),
            g["delta_only"],
            g["drift_only"],
            g["both"],
            g["pct_delta_only_of_selected"],
            g["pct_drift_only_of_selected"],
            g["pct_both_of_selected"],
        ]
    )
    return headers, rows


def build_stage4_threshold() -> tuple[list[str], list[list]]:
    paths = [
        ("MSAD", "1_MSAD_Generate_All_Normal_Window_output_All_1Sigma_s16_Delta_Drift_W20_output/selected_frames_threshold_0.4_count.json"),
        ("UCF-Crime", "1_UCF_Generate_All_Normal_Window_output_All_1Sigma_s16_Delta_Drift_W20_output/selected_frames_threshold_0.4_count.json"),
        ("ShanghaiTech", "1_ShanghiTech_Generate_All_Normal_Window_output_All_1Sigma_s16_Delta_Drift_clip_W20_output_Forward_16/selected_frames_threshold_0.4_count.json"),
    ]
    headers = [
        "Dataset",
        "Threshold τ_f",
        "Stage-2 Selected",
        "Above τ_f (Stage 4)",
        "At/Below τ_f",
        "Out of Bounds",
        "% Above τ_f",
        "% At/Below τ_f",
    ]
    rows = []
    for ds, rel in paths:
        d = load_json(rel)
        rows.append(
            [
                ds,
                d["threshold"],
                d["total_selected_frames"],
                d["above_threshold"],
                d["at_or_below_threshold"],
                d.get("out_of_bounds", 0),
                round(d["pct_above_threshold"], 2),
                round(d["pct_at_or_below_threshold"], 2),
            ]
        )
    # XD from paper table (no stored JSON in repo)
    rows.append([ "XD-Violence", 0.4, 102066, 15968, 86098, "", round(100 * 15968 / 102066, 2), round(100 * 86098 / 102066, 2) ])
    return headers, rows


def build_computational_cost() -> tuple[list[str], list[list]]:
    headers = [
        "Stage",
        "Component",
        "Batch Size (B)",
        "Time T (s)",
        "XD Frames",
        "UCF Frames",
        "MSAD Frames",
        "Shanghai Frames",
        "Notes",
    ]
    rows = [
        [
            "Stage 1",
            "CLIP ViT-B/32",
            24,
            0.14,
            146449,
            69634,
            14360,
            2550,
            "Stride-sampled candidate frames",
        ],
        [
            "Stage 2–3",
            "VLM (Qwen3-VL-8B)",
            24,
            8.19,
            102066,
            53283,
            9353,
            1898,
            "Selected anchor frames",
        ],
        [
            "Stage 2–3",
            "LLM (Qwen2.5-7B)",
            24,
            0.64,
            102066,
            53283,
            9353,
            1898,
            "Selected anchor frames",
        ],
        [
            "Stage 4",
            "VideoLLaMA3-7B",
            1,
            0.84,
            15968,
            3300,
            2327,
            487,
            "Gated anchors (s_frame > τ_f); Shanghai paper=487, computed=580",
        ],
        [
            "Stage 4",
            "LLM (Qwen2.5-7B)",
            24,
            0.64,
            15968,
            3300,
            2327,
            487,
            "Gated anchors",
        ],
        ["Summary", "Total (h)", "", "", 14.4, 6.35, 1.54, 0.33, "End-to-end wall time"],
        ["Summary", "Amortized FPS", "", "", 44.95, 48.61, 41.06, 33.55, "Full-resolution throughput"],
    ]
    return headers, rows


def build_invoke_rate() -> tuple[list[str], list[list]]:
    data = load_json("ablation_invoke_rate/invoke_rate_results.json")
    headers = [
        "Dataset",
        "Threshold τ_f",
        "AUC-ROC (%)",
        "AP (%)",
        "Invoked Frames",
        "Total Frames",
        "Invoke Rate (%)",
    ]
    rows = []
    for ds, entries in data["results"].items():
        for e in entries:
            rows.append(
                [
                    ds,
                    e["threshold_tf"],
                    pct(e["auc_roc"]),
                    pct(e["ap"]),
                    e["invoked_frames"],
                    e["total_frames"],
                    pct(e["invoke_rate"]),
                ]
            )
    return headers, rows


def build_latency() -> tuple[list[str], list[list]]:
    data = load_json("ablation_saaf_breakdown/latency_benchmark.json")
    headers = [
        "Dataset",
        "Videos",
        "Total Frames",
        "Avg Frames/Video",
        "Stage1 ms/video",
        "Stage2 ms/video",
        "Stage3 ms/video",
        "Total ms/video",
        "Throughput FPS",
        "Wall Time (s)",
        "Refine σ",
    ]
    rows = []
    for ds, d in data.items():
        pv = d["per_video_ms"]
        rows.append(
            [
                ds,
                d["n_videos"],
                d["total_frames"],
                d["avg_frames_per_video"],
                pv["stage1_mean"],
                pv["stage2_mean"],
                pv["stage3_mean"],
                pv["total_mean"],
                d["throughput_fps"],
                d["wall_time_sec"],
                d["refine_sigma"],
            ]
        )
    return headers, rows


def save_workbook(path: Path, sheets: list[tuple[str, list[str], list[list]]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for title, headers, rows in sheets:
        ws = wb.create_sheet(title[:31])
        write_table(ws, f"SAAF — {title}", headers, rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def save_multi_table_workbook(path: Path, sheet_title: str, tables: list[tuple[str, list[str], list[list]]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    row = 1
    for title, headers, rows in tables:
        row = write_table(ws, title, headers, rows, start_row=row)
        row += 2
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    out_dir = ROOT / "Excel_Reports"
    out_dir.mkdir(exist_ok=True)

    # 1) Master summary workbook
    var_summary_h, var_summary_r, var_detail_h, var_detail_r = build_variance_seeds()
    master_sheets = [
        ("Final AUC-ROC (Ours)", *build_final_aucroc()),
        ("Stage Breakdown", *build_stage_breakdown()),
        ("Fusion Weight Ablation", *build_fusion_ablation()),
        ("Qwen Seed Variance", var_summary_h, var_summary_r),
        ("Signal Attribution", *build_signal_attribution()),
        ("Stage 4 Gate (τ_f=0.4)", *build_stage4_threshold()),
        ("Computational Cost", *build_computational_cost()),
        ("Invoke Rate Ablation", *build_invoke_rate()),
        ("Latency Benchmark", *build_latency()),
    ]
    save_workbook(out_dir / "SAAF_Results_Summary.xlsx", master_sheets)

    # 2) Individual focused workbooks
    save_workbook(out_dir / "01_Final_AUCROC.xlsx", [("Final AUC-ROC", *build_final_aucroc())])
    save_workbook(out_dir / "02_Stage_Breakdown.xlsx", [("Stage Breakdown", *build_stage_breakdown())])
    save_workbook(out_dir / "03_Fusion_Weight_Ablation.xlsx", [("Fusion Weights", *build_fusion_ablation())])
    save_multi_table_workbook(
        out_dir / "04_Qwen_Variance_Seeds.xlsx",
        "Variance",
        [
            ("Summary (3 seeds)", var_summary_h, var_summary_r),
            ("Per-seed detail", var_detail_h, var_detail_r),
        ],
    )
    save_workbook(out_dir / "05_Signal_Attribution.xlsx", [("Attribution", *build_signal_attribution())])
    save_workbook(out_dir / "06_Stage4_Threshold_0.4.xlsx", [("Stage 4 Gating", *build_stage4_threshold())])
    save_workbook(out_dir / "07_Computational_Cost.xlsx", [("Computational Cost", *build_computational_cost())])

    print(f"Wrote Excel reports to: {out_dir}")
    for p in sorted(out_dir.glob("*.xlsx")):
        print(f"  - {p.name} ({p.stat().st_size // 1024} KB)")


if __name__ == "__main__":
    main()
